from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
import io
import re
import time
import uuid
from typing import Any
from bisect import bisect_right

from flask import Blueprint, current_app, jsonify, request, send_file
from flask_jwt_extended import get_jwt, get_jwt_identity, jwt_required
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, insert, or_, text
from sqlalchemy.exc import IntegrityError, SQLAlchemyError

from extensions import db
from models import (
    Company,
    ExsolInventoryItem,
    ExsolProductionEntry,
    ExsolProductionSerial,
    ExsolSerialEvent,
    RoleEnum,
    User,
    normalize_role,
)


bp = Blueprint("exsol_production", __name__, url_prefix="/api/exsol/production")

EXSOL_COMPANY_KEY = "EXSOL"
EXSOL_COMPANY_NAME = "Exsol Engineering (Pvt) Ltd"
SHIFT_OPTIONS = {"Morning", "Evening", "Night"}
SERIAL_REGEX = re.compile(r"^[0-9]{8}$")
MAX_BULK_QUANTITY = 500
MAX_BULK_CHUNK_SIZE = 50
SERIAL_CHUNK_SIZE = 200
_EXSOL_SEQUENCES_READY = False

def _build_error(message: str, status: int = 400, details: list[str] | None = None):
    payload: dict[str, Any] = {"ok": False, "error": message}
    if details:
        payload["details"] = details
    return jsonify(payload), status


def _has_exsol_production_access() -> bool:
    try:
        claims = get_jwt()
    except Exception:
        return False

    role_raw = claims.get("role")
    company_key = (claims.get("company_key") or claims.get("company") or "").strip().lower()
    role = normalize_role(role_raw)

    if role not in {RoleEnum.sales_manager, RoleEnum.sales_executive, RoleEnum.admin}:
        return False

    if role != RoleEnum.admin and company_key and company_key != "exsol-engineering":
        return False

    return True


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None
    return None


def _parse_quantity(value: Any) -> int | None:
    try:
        quantity = int(value)
    except (TypeError, ValueError):
        return None
    return quantity if quantity > 0 else None


def _split_serials(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        serials: list[str] = []
        for entry in value:
            serials.extend(_split_serials(entry))
        return serials
    text = str(value).strip()
    if not text:
        return []
    parts = re.split(r"[\s,]+", text)
    return [part.strip() for part in parts if part and part.strip()]


def _normalize_serial(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        value = str(int(value)).zfill(8)
    text = str(value).strip()
    if not text:
        return None
    if SERIAL_REGEX.match(text):
        return text
    return None


def _normalize_serial_mode(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in {"serialrange", "serial_range", "range"}:
        return "SerialRange"
    if text in {"manual", "list"}:
        return "Manual"
    return None


def _normalize_uuid(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return str(uuid.UUID(text))
    except (ValueError, AttributeError, TypeError):
        return None


def _serialize_entry(entry: ExsolProductionEntry, serial: ExsolProductionSerial, user_lookup: dict[int, str]):
    return {
        "id": entry.id,
        "company_key": entry.company_key,
        "production_date": entry.production_date.isoformat(),
        "item_code": entry.item_code,
        "item_name": entry.item_name,
        "serial_number": serial.serial_no,
        "production_shift": entry.shift,
        "created_by": entry.created_by_user_id,
        "created_by_name": entry.created_by_name or user_lookup.get(entry.created_by_user_id, "Unknown"),
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "is_confirmed": bool(entry.is_confirmed),
        "confirmed_by": entry.confirmed_by_user_id,
        "confirmed_at": entry.confirmed_at.isoformat() if entry.confirmed_at else None,
    }


def _load_user_lookup(user_ids: set[int]) -> dict[int, str]:
    if not user_ids:
        return {}
    users = User.query.filter(User.id.in_(user_ids)).all()
    return {user.id: user.name for user in users}


def _ensure_exsol_sequences() -> None:
    global _EXSOL_SEQUENCES_READY
    if _EXSOL_SEQUENCES_READY:
        return
    bind = db.session.get_bind()
    if not bind or bind.dialect.name != "postgresql":
        return
    entries_seq = "exsol_production_entries_id_seq"
    serials_seq = "exsol_production_serials_id_seq"
    try:
        with bind.begin() as connection:
            connection.execute(text(f"CREATE SEQUENCE IF NOT EXISTS {entries_seq}"))
            connection.execute(text(f"CREATE SEQUENCE IF NOT EXISTS {serials_seq}"))
            connection.execute(
                text(
                    "ALTER TABLE exsol_production_entries "
                    f"ALTER COLUMN id SET DEFAULT nextval('{entries_seq}')"
                )
            )
            connection.execute(
                text(
                    "ALTER TABLE exsol_production_serials "
                    f"ALTER COLUMN id SET DEFAULT nextval('{serials_seq}')"
                )
            )
            connection.execute(
                text(
                    "SELECT setval("
                    f"'{entries_seq}', "
                    "GREATEST(COALESCE((SELECT MAX(id) FROM exsol_production_entries), 0), 1), "
                    "true)"
                )
            )
            connection.execute(
                text(
                    "SELECT setval("
                    f"'{serials_seq}', "
                    "GREATEST(COALESCE((SELECT MAX(id) FROM exsol_production_serials), 0), 1), "
                    "true)"
                )
            )
        _EXSOL_SEQUENCES_READY = True
    except SQLAlchemyError as exc:
        current_app.logger.warning(
            {
                "event": "exsol_sequence_setup_failed",
                "message": str(exc),
            }
        )


def _get_exsol_company_id() -> int | None:
    company = Company.query.filter(Company.name == EXSOL_COMPANY_NAME).one_or_none()
    return company.id if company else None


def _lookup_exsol_item(company_id: int, item_id: str | None, item_code: str | None) -> ExsolInventoryItem | None:
    if item_id:
        return (
            ExsolInventoryItem.query.filter(
                ExsolInventoryItem.company_id == company_id,
                ExsolInventoryItem.id == item_id,
                ExsolInventoryItem.is_active.is_(True),
            )
            .one_or_none()
        )
    if item_code:
        return (
            ExsolInventoryItem.query.filter(
                ExsolInventoryItem.company_id == company_id,
                ExsolInventoryItem.is_active.is_(True),
                func.lower(ExsolInventoryItem.item_code) == item_code.lower(),
            )
            .one_or_none()
        )
    return None


def _summarize_payload(rows: list[dict[str, Any]]) -> dict[str, Any]:
    quantities = []
    serial_counts = []
    serial_min = None
    serial_max = None
    for row in rows:
        quantity = _parse_quantity(row.get("quantity"))
        if quantity:
            quantities.append(quantity)
        serial_mode = _normalize_serial_mode(row.get("serial_mode"))
        if serial_mode == "Manual":
            serials = row.get("serials") or row.get("serial_numbers")
            serial_list = [_normalize_serial(value) for value in _split_serials(serials)]
            serial_list = [value for value in serial_list if value]
            if serial_list:
                serial_min = min(serial_min or serial_list[0], min(serial_list))
                serial_max = max(serial_max or serial_list[0], max(serial_list))
            serial_counts.append(len(serial_list))
        elif serial_mode == "SerialRange":
            start_serial = _normalize_serial(row.get("start_serial") or row.get("starting_serial"))
            if start_serial and quantity:
                start_int = int(start_serial)
                end_int = start_int + quantity - 1
                if end_int <= 99999999:
                    first = str(start_int).zfill(8)
                    last = str(end_int).zfill(8)
                    serial_min = min(serial_min or first, first)
                    serial_max = max(serial_max or last, last)
                    serial_counts.append(quantity)
    summary = {
        "rows": len(rows),
        "quantity_total": sum(quantities),
        "serial_count_total": sum(serial_counts),
    }
    if serial_min and serial_max:
        summary["serial_summary"] = {"count": sum(serial_counts), "first": serial_min, "last": serial_max}
    return summary


def _chunk_rows(rows: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [rows[i : i + size] for i in range(0, len(rows), size)]


def _chunk_values(values: list[Any], size: int) -> list[list[Any]]:
    return [values[i : i + size] for i in range(0, len(values), size)]


def _log_bulk_event(event: str, payload: dict[str, Any]) -> None:
    current_app.logger.info({"event": event, **payload})


def _log_bulk_error(payload: dict[str, Any]) -> None:
    current_app.logger.error({"event": "exsol_production_bulk_error", **payload})


def _prepare_bulk_chunk(
    rows: list[dict[str, Any]],
    base_index: int,
    company_id: int,
    user_id: int,
    role_name: str,
    user_name: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    errors: dict[int, list[dict[str, str]]] = defaultdict(list)

    item_ids = {
        normalized
        for row in rows
        for normalized in (_normalize_uuid(row.get("item_id")),)
        if normalized
    }
    item_codes = {
        (row.get("item_code") or "").strip()
        for row in rows
        if (row.get("item_code") or "").strip()
    }
    filters = []
    if item_ids:
        filters.append(ExsolInventoryItem.id.in_(item_ids))
    if item_codes:
        filters.append(func.lower(ExsolInventoryItem.item_code).in_({code.lower() for code in item_codes}))
    items: list[ExsolInventoryItem] = []
    if filters:
        items = (
            ExsolInventoryItem.query.filter(
                ExsolInventoryItem.company_id == company_id,
                ExsolInventoryItem.is_active.is_(True),
            )
            .filter(or_(*filters))
            .all()
        )
    item_lookup_by_code = {item.item_code.lower(): item for item in items}
    item_lookup_by_id = {str(item.id): item for item in items}

    entry_rows: list[dict[str, Any]] = []
    row_specs: list[dict[str, Any]] = []
    for offset, row in enumerate(rows):
        idx = base_index + offset
        production_date = _parse_date(row.get("production_date"))
        if not production_date:
            errors[idx].append({"row": idx, "field": "production_date", "message": "Production date is required."})

        item_id_raw = (row.get("item_id") or "").strip()
        item_id = _normalize_uuid(item_id_raw) or ""
        item_code = (row.get("item_code") or "").strip()
        if not item_id and not item_code:
            errors[idx].append({"row": idx, "field": "item_code", "message": "Item code is required."})
            item = None
        else:
            item = item_lookup_by_id.get(item_id) if item_id else item_lookup_by_code.get(item_code.lower())
            if not item:
                label = item_code or item_id
                errors[idx].append(
                    {"row": idx, "field": "item_code", "message": f"Item {label} is not an Exsol inventory item."}
                )

        shift = (row.get("shift") or row.get("production_shift") or "").strip()
        if shift and shift not in SHIFT_OPTIONS:
            errors[idx].append(
                {
                    "row": idx,
                    "field": "production_shift",
                    "message": "Production shift must be Morning, Evening, or Night.",
                }
            )

        quantity = _parse_quantity(row.get("quantity"))
        if quantity is not None and quantity > MAX_BULK_QUANTITY:
            errors[idx].append(
                {"row": idx, "field": "quantity", "message": f"Quantity cannot exceed {MAX_BULK_QUANTITY}."}
            )

        serial_mode = _normalize_serial_mode(row.get("serial_mode"))
        if not serial_mode:
            errors[idx].append(
                {"row": idx, "field": "serial_mode", "message": "Serial mode must be SerialRange or Manual."}
            )

        serials: list[str] = []
        start_int = None
        end_int = None
        start_serial = None
        end_serial = None
        if serial_mode == "Manual":
            raw_serials = row.get("serials") or row.get("serial_numbers")
            serials = [_normalize_serial(value) or "" for value in _split_serials(raw_serials)]
            invalid_serials = [value for value in serials if not SERIAL_REGEX.match(value)]
            serials = [value for value in serials if SERIAL_REGEX.match(value)]
            if invalid_serials:
                errors[idx].append(
                    {"row": idx, "field": "serials", "message": "All serial numbers must be exactly 8 digits."}
                )
            if not serials:
                errors[idx].append(
                    {"row": idx, "field": "serials", "message": "Serial numbers are required for manual entry."}
                )
            if quantity is None:
                quantity = len(serials)
            elif quantity and len(serials) != quantity:
                errors[idx].append(
                    {"row": idx, "field": "quantity", "message": "Quantity must match the number of serials provided."}
                )
            if quantity and quantity > MAX_BULK_QUANTITY:
                errors[idx].append(
                    {"row": idx, "field": "quantity", "message": f"Quantity cannot exceed {MAX_BULK_QUANTITY}."}
                )
        elif serial_mode == "SerialRange":
            starting_serial = row.get("start_serial") or row.get("starting_serial")
            if not starting_serial:
                errors[idx].append(
                    {"row": idx, "field": "serial_start", "message": "Starting serial is required for serial range."}
                )
            if quantity is None:
                errors[idx].append(
                    {"row": idx, "field": "quantity", "message": "Quantity must be provided to generate serials."}
                )
            else:
                normalized = _normalize_serial(str(starting_serial or ""))
                if not normalized:
                    errors[idx].append(
                        {"row": idx, "field": "serial_start", "message": "Starting serial must be exactly 8 digits."}
                    )
                else:
                    start_int = int(normalized)
                    end_int = start_int + quantity - 1
                    if end_int > 99999999:
                        errors[idx].append(
                            {
                                "row": idx,
                                "field": "serial_start",
                                "message": "Serial range exceeds 8 digits. Adjust the starting serial or quantity.",
                            }
                        )
                    else:
                        start_serial = str(start_int).zfill(8)
                        end_serial = str(end_int).zfill(8)

        if quantity is None:
            errors[idx].append(
                {"row": idx, "field": "quantity", "message": "Quantity must be a positive integer."}
            )

        if errors[idx]:
            continue

        if serial_mode == "Manual":
            seen_serials: set[str] = set()
            for serial in serials:
                if serial in seen_serials:
                    errors[idx].append(
                        {"row": idx, "field": "serials", "message": f"Serial {serial} is duplicated in this row."}
                    )
                seen_serials.add(serial)

        row_specs.append(
            {
                "row_index": idx,
                "item": item,
                "item_code": item.item_code if item else item_code,
                "item_name": item.item_name if item else (row.get("item_name") or item_code),
                "production_date": production_date,
                "shift": shift or None,
                "quantity": quantity,
                "serial_mode": serial_mode,
                "serials": serials,
                "start_int": start_int,
                "end_int": end_int,
                "start_serial": start_serial,
                "end_serial": end_serial,
            }
        )

        entry_rows.append(
            {
                "company_key": EXSOL_COMPANY_KEY,
                "production_date": production_date,
                "item_code": item.item_code if item else item_code,
                "item_name": item.item_name if item else (row.get("item_name") or item_code),
                "shift": shift or None,
                "quantity": quantity,
                "serial_mode": serial_mode,
                "created_by_user_id": user_id,
                "created_by_name": user_name or None,
            }
        )

    error_list = []
    for idx, messages in errors.items():
        error_list.extend(messages)
    return entry_rows, row_specs, error_list


def _find_batch_duplicate_errors(
    row_specs: list[dict[str, Any]],
) -> tuple[list[dict[str, str]], list[str]]:
    errors: list[dict[str, str]] = []
    duplicate_serials: list[str] = []
    manual_serials: dict[str, int] = {}
    manual_serial_values: list[tuple[str, int]] = []
    ranges: list[tuple[int, int, int]] = []

    for spec in row_specs:
        idx = spec["row_index"]
        if spec["serial_mode"] == "Manual":
            for serial in spec["serials"]:
                if serial in manual_serials:
                    errors.append(
                        {"row": idx, "field": "serials", "message": f"Serial {serial} is duplicated in this batch."}
                    )
                    errors.append(
                        {
                            "row": manual_serials[serial],
                            "field": "serials",
                            "message": f"Serial {serial} is duplicated in this batch.",
                        }
                    )
                    duplicate_serials.append(serial)
                manual_serials[serial] = idx
                manual_serial_values.append((serial, idx))
        elif spec["serial_mode"] == "SerialRange" and spec["start_int"] is not None and spec["end_int"] is not None:
            ranges.append((spec["start_int"], spec["end_int"], idx))

    if ranges:
        ranges_sorted = sorted(ranges, key=lambda value: value[0])
        prev_start, prev_end, prev_idx = ranges_sorted[0]
        for start, end, idx in ranges_sorted[1:]:
            if start <= prev_end:
                errors.append(
                    {
                        "row": idx,
                        "field": "serial_start",
                        "message": f"Serial range overlaps existing range starting {str(prev_start).zfill(8)}.",
                    }
                )
                errors.append(
                    {
                        "row": prev_idx,
                        "field": "serial_start",
                        "message": f"Serial range overlaps existing range starting {str(start).zfill(8)}.",
                    }
                )
                duplicate_serials.append(str(start).zfill(8))
            if end > prev_end:
                prev_start, prev_end, prev_idx = start, end, idx

    if ranges and manual_serial_values:
        ranges_sorted = sorted(ranges, key=lambda value: value[0])
        range_starts = [start for start, _, _ in ranges_sorted]
        for serial, idx in manual_serial_values:
            serial_int = int(serial)
            position = bisect_right(range_starts, serial_int) - 1
            if position >= 0:
                range_start, range_end, range_idx = ranges_sorted[position]
                if range_start <= serial_int <= range_end:
                    errors.append(
                        {
                            "row": idx,
                            "field": "serials",
                            "message": f"Serial {serial} is duplicated in this batch.",
                        }
                    )
                    errors.append(
                        {
                            "row": range_idx,
                            "field": "serial_start",
                            "message": f"Serial {serial} is duplicated in this batch.",
                        }
                    )
                    duplicate_serials.append(serial)

    return errors, duplicate_serials


def _find_cross_chunk_duplicate_errors(
    row_specs: list[dict[str, Any]],
    seen_serials: set[str],
    seen_ranges: list[tuple[int, int]],
) -> tuple[list[dict[str, str]], list[str], set[str], list[tuple[int, int]]]:
    errors: list[dict[str, str]] = []
    duplicate_serials: list[str] = []
    new_serials: set[str] = set()
    new_ranges: list[tuple[int, int]] = []

    def _serial_in_seen_ranges(serial_int: int) -> bool:
        return any(start <= serial_int <= end for start, end in seen_ranges)

    for spec in row_specs:
        idx = spec["row_index"]
        if spec["serial_mode"] == "Manual":
            for serial in spec["serials"]:
                serial_int = int(serial)
                if serial in seen_serials or _serial_in_seen_ranges(serial_int):
                    errors.append(
                        {"row": idx, "field": "serials", "message": f"Serial {serial} is duplicated in this batch."}
                    )
                    duplicate_serials.append(serial)
                new_serials.add(serial)
        elif spec["serial_mode"] == "SerialRange" and spec["start_int"] is not None and spec["end_int"] is not None:
            start_int = spec["start_int"]
            end_int = spec["end_int"]
            overlaps_range = next(
                ((start, end) for start, end in seen_ranges if start_int <= end and end_int >= start),
                None,
            )
            if overlaps_range:
                errors.append(
                    {
                        "row": idx,
                        "field": "serial_start",
                        "message": "Serial range overlaps another range in this batch.",
                    }
                )
                duplicate_serials.append(str(max(start_int, overlaps_range[0])).zfill(8))
            if not overlaps_range and any(
                start_int <= int(serial) <= end_int for serial in seen_serials
            ):
                candidate = next(
                    (serial for serial in seen_serials if start_int <= int(serial) <= end_int),
                    str(start_int).zfill(8),
                )
                errors.append(
                    {
                        "row": idx,
                        "field": "serial_start",
                        "message": f"Serial {candidate} is duplicated in this batch.",
                    }
                )
                duplicate_serials.append(candidate)
            new_ranges.append((start_int, end_int))

    return errors, duplicate_serials, new_serials, new_ranges


def _find_existing_serials(
    row_specs: list[dict[str, Any]],
) -> set[str]:
    manual_serials: list[str] = []
    range_specs: list[tuple[str, str]] = []
    for spec in row_specs:
        if spec["serial_mode"] == "Manual":
            manual_serials.extend(spec["serials"])
        elif spec["serial_mode"] == "SerialRange" and spec["start_serial"] and spec["end_serial"]:
            range_specs.append((spec["start_serial"], spec["end_serial"]))

    if not manual_serials and not range_specs:
        return set()

    existing_serials: set[str] = set()
    if manual_serials:
        for chunk in _chunk_values(manual_serials, SERIAL_CHUNK_SIZE):
            rows = (
                db.session.query(ExsolProductionSerial.serial_no)
                .filter(ExsolProductionSerial.company_key == EXSOL_COMPANY_KEY)
                .filter(ExsolProductionSerial.serial_no.in_(chunk))
                .all()
            )
            existing_serials.update(serial_no for (serial_no,) in rows)

    if range_specs:
        for chunk in _chunk_values(range_specs, SERIAL_CHUNK_SIZE):
            filters = [
                ExsolProductionSerial.serial_no.between(start_serial, end_serial)
                for start_serial, end_serial in chunk
            ]
            rows = (
                db.session.query(ExsolProductionSerial.serial_no)
                .filter(ExsolProductionSerial.company_key == EXSOL_COMPANY_KEY)
                .filter(or_(*filters))
                .all()
            )
            existing_serials.update(serial_no for (serial_no,) in rows)

    return existing_serials


def _check_existing_serial_errors(
    row_specs: list[dict[str, Any]],
    existing_serials: set[str],
) -> tuple[list[dict[str, str]], list[str]]:
    errors: list[dict[str, str]] = []
    duplicate_serials: list[str] = []
    existing_serials_sorted = sorted(existing_serials)
    for spec in row_specs:
        idx = spec["row_index"]
        if spec["serial_mode"] == "Manual":
            conflicts = [serial for serial in spec["serials"] if serial in existing_serials]
            for serial in conflicts:
                errors.append(
                    {
                        "row": idx,
                        "field": "serial_start",
                        "message": f"Serial already exists: {serial}",
                    }
                )
                duplicate_serials.append(serial)
        elif spec["serial_mode"] == "SerialRange" and spec["start_serial"] and spec["end_serial"]:
            if existing_serials_sorted:
                position = bisect_right(existing_serials_sorted, spec["end_serial"]) - 1
                if position >= 0:
                    candidate = existing_serials_sorted[position]
                    if spec["start_serial"] <= candidate <= spec["end_serial"]:
                        errors.append(
                            {
                                "row": idx,
                                "field": "serial_start",
                                "message": f"Serial already exists: {candidate}",
                            }
                        )
                        duplicate_serials.append(candidate)
    return errors, duplicate_serials


def _insert_bulk_chunk(
    entry_rows: list[dict[str, Any]],
    row_specs: list[dict[str, Any]],
) -> tuple[int, int]:
    if not entry_rows:
        return 0, 0

    now = datetime.utcnow()
    for offset, row in enumerate(entry_rows):
        row.setdefault("created_at", now + timedelta(microseconds=offset))

    _ensure_exsol_sequences()
    bind = db.session.get_bind()
    dialect_name = bind.dialect.name if bind else ""
    entry_ids: list[int] = []
    if dialect_name == "postgresql":
        insert_stmt = insert(ExsolProductionEntry).returning(ExsolProductionEntry.id)
        result = db.session.execute(insert_stmt, entry_rows)
        entry_ids = [row_id for row_id in result.scalars().all()]
    else:
        max_id = db.session.query(func.max(ExsolProductionEntry.id)).scalar() or 0
        for offset, row in enumerate(entry_rows, start=1):
            row["id"] = max_id + offset
        insert_stmt = insert(ExsolProductionEntry)
        db.session.execute(insert_stmt, entry_rows)
        entry_ids = [row["id"] for row in entry_rows]

    total_serials = 0
    serial_insert_rows: list[dict[str, Any]] = []
    event_insert_rows: list[dict[str, Any]] = []
    bind = db.session.get_bind()
    dialect_name = bind.dialect.name if bind else ""

    event_created_at = datetime.utcnow()
    for entry_id, spec, entry_row in zip(entry_ids, row_specs, entry_rows):
        event_date = entry_row.get("created_at") or event_created_at
        if spec["serial_mode"] == "SerialRange" and spec["start_int"] is not None and spec["end_int"] is not None:
            total_serials += spec["quantity"]
            if dialect_name == "postgresql":
                db.session.execute(
                    text(
                        "INSERT INTO exsol_production_serials "
                        "(company_key, serial_no, entry_id, created_at) "
                        "SELECT :company_key, to_char(gs, 'FM00000000'), :entry_id, now() "
                        "FROM generate_series("
                        "CAST(:start_serial AS bigint), CAST(:end_serial AS bigint)"
                        ") gs"
                    ),
                    {
                        "company_key": EXSOL_COMPANY_KEY,
                        "entry_id": entry_id,
                        "start_serial": spec["start_int"],
                        "end_serial": spec["end_int"],
                    },
                )
            else:
                for serial_value in range(spec["start_int"], spec["end_int"] + 1):
                    serial_insert_rows.append(
                        {
                            "company_key": EXSOL_COMPANY_KEY,
                            "serial_no": str(serial_value).zfill(8),
                            "entry_id": entry_id,
                            "created_at": datetime.utcnow(),
                        }
                    )
            for serial_value in range(spec["start_int"], spec["end_int"] + 1):
                event_insert_rows.append(
                    {
                        "company_key": EXSOL_COMPANY_KEY,
                        "item_code": spec["item_code"],
                        "serial_number": str(serial_value).zfill(8),
                        "event_type": "PRODUCED",
                        "event_date": event_date,
                        "ref_type": "PRODUCTION_ENTRY",
                        "ref_id": str(entry_id),
                        "created_at": event_created_at,
                    }
                )
        else:
            for serial in spec["serials"]:
                serial_insert_rows.append(
                    {
                        "company_key": EXSOL_COMPANY_KEY,
                        "serial_no": serial,
                        "entry_id": entry_id,
                        "created_at": datetime.utcnow(),
                    }
                )
                event_insert_rows.append(
                    {
                        "company_key": EXSOL_COMPANY_KEY,
                        "item_code": spec["item_code"],
                        "serial_number": serial,
                        "event_type": "PRODUCED",
                        "event_date": event_date,
                        "ref_type": "PRODUCTION_ENTRY",
                        "ref_id": str(entry_id),
                        "created_at": event_created_at,
                    }
                )
            total_serials += len(spec["serials"])

    if serial_insert_rows:
        serial_insert_stmt = insert(ExsolProductionSerial)
        if dialect_name != "postgresql":
            max_serial_id = db.session.query(func.max(ExsolProductionSerial.id)).scalar() or 0
            serial_id_offset = max_serial_id
            for chunk in _chunk_values(serial_insert_rows, SERIAL_CHUNK_SIZE):
                for offset, row in enumerate(chunk, start=1):
                    row["id"] = serial_id_offset + offset
                serial_id_offset += len(chunk)
                db.session.execute(serial_insert_stmt, chunk)
        else:
            for chunk in _chunk_values(serial_insert_rows, SERIAL_CHUNK_SIZE):
                db.session.execute(serial_insert_stmt, chunk)

    if event_insert_rows:
        event_insert_stmt = insert(ExsolSerialEvent)
        for chunk in _chunk_values(event_insert_rows, SERIAL_CHUNK_SIZE):
            db.session.execute(event_insert_stmt, chunk)

    return len(entry_ids), total_serials


@bp.get("/entries")
@jwt_required()
def list_entries():
    if not _has_exsol_production_access():
        return _build_error("Access denied", 403)

    params = request.args
    start_date = _parse_date(params.get("date_from") or params.get("start_date"))
    end_date = _parse_date(params.get("date_to") or params.get("end_date"))
    item_code = (params.get("item_code") or "").strip()
    serial_no = (params.get("serial_no") or "").strip()
    production_shift = (params.get("production_shift") or "").strip()
    created_by = params.get("created_by")
    confirmed = (params.get("confirmed") or "").strip().lower()
    try:
        limit = int(params.get("limit", 250))
    except (TypeError, ValueError):
        return _build_error("Invalid limit parameter.", 400)
    limit = min(max(limit, 1), 500)

    query = (
        db.session.query(ExsolProductionEntry, ExsolProductionSerial)
        .join(ExsolProductionSerial, ExsolProductionSerial.entry_id == ExsolProductionEntry.id)
        .filter(ExsolProductionEntry.company_key == EXSOL_COMPANY_KEY)
    )

    if start_date:
        query = query.filter(ExsolProductionEntry.production_date >= start_date)
    if end_date:
        query = query.filter(ExsolProductionEntry.production_date <= end_date)
    if item_code:
        query = query.filter(func.lower(ExsolProductionEntry.item_code) == item_code.lower())
    if serial_no:
        query = query.filter(ExsolProductionSerial.serial_no == serial_no)
    if production_shift:
        query = query.filter(ExsolProductionEntry.shift == production_shift)
    if created_by:
        try:
            created_id = int(created_by)
            query = query.filter(ExsolProductionEntry.created_by_user_id == created_id)
        except (TypeError, ValueError):
            return _build_error("Invalid created_by filter", 400)
    if confirmed in {"true", "false"}:
        query = query.filter(ExsolProductionEntry.is_confirmed == (confirmed == "true"))

    rows = (
        query.order_by(
            ExsolProductionEntry.production_date.desc(),
            ExsolProductionEntry.created_at.desc(),
            ExsolProductionSerial.serial_no.asc(),
        )
        .limit(limit)
        .all()
    )
    user_lookup = _load_user_lookup({entry.created_by_user_id for entry, _ in rows})
    return jsonify([_serialize_entry(entry, serial, user_lookup) for entry, serial in rows])


@bp.post("/bulk")
@jwt_required()
def bulk_create_entries():
    if not _has_exsol_production_access():
        return _build_error("Access denied", 403)

    start_time = time.monotonic()
    request_id = str(uuid.uuid4())
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows") or []
    if not isinstance(rows, list) or not rows:
        return _build_error("Rows payload is required.", 400)

    user_id_raw = get_jwt_identity()
    try:
        user_id = int(user_id_raw)
    except (TypeError, ValueError):
        return _build_error("Invalid user identity.", 403)

    claims = get_jwt()
    role_name = (claims.get("role") or "").strip()
    user = User.query.get(user_id)
    user_name = user.name if user else ""

    company_id = _get_exsol_company_id()
    if not company_id:
        return _build_error("Exsol company is not configured.", 500)

    _log_bulk_event(
        "exsol_production_bulk_start",
        {
            "user_id": user_id,
            "role": role_name,
            "payload": _summarize_payload(rows),
            "request_id": request_id,
        },
    )

    inserted_rows = 0
    inserted_serials = 0
    total_quantity = 0
    errors: list[dict[str, Any]] = []
    duplicate_serials: list[str] = []
    validation_ms = 0
    db_check_ms = 0
    insert_ms = 0
    commit_ms = 0

    prepared_chunks: list[tuple[list[dict[str, Any]], list[dict[str, Any]]]] = []
    seen_serials: set[str] = set()
    seen_ranges: list[tuple[int, int]] = []

    try:
        for chunk_index, chunk in enumerate(_chunk_rows(rows, MAX_BULK_CHUNK_SIZE)):
            base_index = chunk_index * MAX_BULK_CHUNK_SIZE
            validation_start = time.monotonic()
            entry_rows, row_specs, chunk_errors = _prepare_bulk_chunk(
                chunk,
                base_index,
                company_id,
                user_id,
                role_name,
                user_name,
            )
            validation_ms += int((time.monotonic() - validation_start) * 1000)
            if chunk_errors:
                errors.extend(chunk_errors)
                continue

            batch_errors, batch_duplicates = _find_batch_duplicate_errors(row_specs)
            if batch_errors:
                errors.extend(batch_errors)
                duplicate_serials.extend(batch_duplicates)
                continue

            cross_errors, cross_duplicates, new_serials, new_ranges = _find_cross_chunk_duplicate_errors(
                row_specs,
                seen_serials,
                seen_ranges,
            )
            if cross_errors:
                errors.extend(cross_errors)
                duplicate_serials.extend(cross_duplicates)
                continue

            db_check_start = time.monotonic()
            existing_serials = _find_existing_serials(row_specs)
            db_check_ms += int((time.monotonic() - db_check_start) * 1000)
            existing_errors, existing_duplicates = _check_existing_serial_errors(row_specs, existing_serials)
            if existing_errors:
                errors.extend(existing_errors)
                duplicate_serials.extend(existing_duplicates)
                continue

            seen_serials.update(new_serials)
            seen_ranges.extend(new_ranges)
            prepared_chunks.append((entry_rows, row_specs))
            total_quantity += sum(spec["quantity"] for spec in row_specs)

        duration_ms = int((time.monotonic() - start_time) * 1000)
        if errors:
            error_payload = {
                "user_id": user_id,
                "role": role_name,
                "payload": _summarize_payload(rows),
                "duration_ms": duration_ms,
                "request_id": request_id,
                "error_type": "validation_failed",
                "timings_ms": {
                    "validation_ms": validation_ms,
                    "db_check_ms": db_check_ms,
                },
            }
            _log_bulk_error(error_payload)
            status = 409 if duplicate_serials else 400
            response_payload = {
                "ok": False,
                "message": (
                    f"Serial already exists: {duplicate_serials[0]}"
                    if duplicate_serials
                    else "validation_failed"
                ),
                "errors": errors,
                "request_id": request_id,
            }
            if duplicate_serials:
                response_payload["duplicate_serial"] = duplicate_serials[0]
            return jsonify(response_payload), status

        db.session.rollback()
        try:
            for entry_rows, row_specs in prepared_chunks:
                chunk_insert_start = time.monotonic()
                chunk_rows_inserted, chunk_serials_inserted = _insert_bulk_chunk(
                    entry_rows,
                    row_specs,
                )
                insert_ms += int((time.monotonic() - chunk_insert_start) * 1000)
                inserted_rows += chunk_rows_inserted
                inserted_serials += chunk_serials_inserted
            commit_start = time.monotonic()
            db.session.commit()
            commit_ms = int((time.monotonic() - commit_start) * 1000)
        except IntegrityError:
            db.session.rollback()
            duplicate_serial = None
            for _, row_specs in prepared_chunks:
                existing_serials = _find_existing_serials(row_specs)
                _, existing_duplicates = _check_existing_serial_errors(row_specs, existing_serials)
                if existing_duplicates:
                    duplicate_serial = existing_duplicates[0]
                    break
            return (
                jsonify(
                    {
                        "ok": False,
                        "message": (
                            f"Serial already exists: {duplicate_serial}"
                            if duplicate_serial
                            else "Serial already exists."
                        ),
                        "duplicate_serial": duplicate_serial,
                        "request_id": request_id,
                    }
                ),
                409,
            )
        except SQLAlchemyError as exc:
            db.session.rollback()
            raise exc

        duration_ms = int((time.monotonic() - start_time) * 1000)
        _log_bulk_event(
            "exsol_production_bulk_done",
            {
                "duration_ms": duration_ms,
                "inserted_counts": {"rows": inserted_rows, "serials": inserted_serials},
                "request_id": request_id,
                "timings_ms": {
                    "validation_ms": validation_ms,
                    "db_check_ms": db_check_ms,
                    "insert_ms": insert_ms,
                    "commit_ms": commit_ms,
                },
            },
        )
        return jsonify(
            {
                "ok": True,
                "inserted_production": inserted_rows,
                "inserted_serials": inserted_serials,
                "total_quantity": total_quantity,
            }
        )
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        duration_ms = int((time.monotonic() - start_time) * 1000)
        _log_bulk_error(
            {
                "user_id": user_id,
                "role": role_name,
                "payload": _summarize_payload(rows),
                "duration_ms": duration_ms,
                "request_id": request_id,
                "error_type": type(exc).__name__,
                "timings_ms": {
                    "validation_ms": validation_ms,
                    "db_check_ms": db_check_ms,
                    "insert_ms": insert_ms,
                    "commit_ms": commit_ms,
                },
            }
        )
        return (
            jsonify({"ok": False, "message": "server_error", "request_id": request_id}),
            500,
        )
    finally:
        db.session.close()


@bp.post("/confirm")
@jwt_required()
def confirm_entries():
    if not _has_exsol_production_access():
        return _build_error("Access denied", 403)

    payload = request.get_json(silent=True) or {}
    entry_ids = payload.get("entry_ids") or []
    if not isinstance(entry_ids, list) or not entry_ids:
        return _build_error("entry_ids is required.", 400)

    user_id_raw = get_jwt_identity()
    try:
        user_id = int(user_id_raw)
    except (TypeError, ValueError):
        return _build_error("Invalid user identity.", 403)

    entries = (
        ExsolProductionEntry.query.filter(ExsolProductionEntry.company_key == EXSOL_COMPANY_KEY)
        .filter(ExsolProductionEntry.id.in_(entry_ids))
        .all()
    )
    if not entries:
        return _build_error("No matching entries found.", 404)

    now = datetime.utcnow()
    updated = 0
    try:
        for entry in entries:
            if entry.is_confirmed:
                continue
            entry.is_confirmed = True
            entry.confirmed_by_user_id = user_id
            entry.confirmed_at = now
            updated += 1
            db.session.add(entry)
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        return _build_error("Unable to confirm entries right now.", 500)

    return jsonify({"ok": True, "confirmed": updated})


@bp.patch("/entries/<int:entry_id>")
@jwt_required()
def update_entry(entry_id: int):
    if not _has_exsol_production_access():
        return _build_error("Access denied", 403)

    payload = request.get_json(silent=True) or {}

    entry = (
        ExsolProductionEntry.query.filter(ExsolProductionEntry.company_key == EXSOL_COMPANY_KEY)
        .filter(ExsolProductionEntry.id == entry_id)
        .one_or_none()
    )
    if not entry:
        return _build_error("Entry not found.", 404)
    if entry.is_confirmed:
        return _build_error("Confirmed entries cannot be edited.", 400)
    if entry.production_date != date.today():
        return _build_error("Entries can only be edited on the production date.", 400)

    updates: dict[str, Any] = {}
    if "production_shift" in payload or "shift" in payload:
        production_shift = (payload.get("shift") or payload.get("production_shift") or "").strip()
        if production_shift and production_shift not in SHIFT_OPTIONS:
            return _build_error("Production shift must be Morning, Evening, or Night.", 400)
        updates["shift"] = production_shift or None

    if "item_code" in payload or "item_id" in payload:
        item_id = (payload.get("item_id") or "").strip()
        item_code = (payload.get("item_code") or "").strip()
        if not item_id and not item_code:
            return _build_error("Item code is required.", 400)
        company_id = _get_exsol_company_id()
        if not company_id:
            return _build_error("Exsol company not configured.", 500)
        item = _lookup_exsol_item(company_id, item_id, item_code)
        if not item:
            return _build_error("Item is not an Exsol inventory item.", 400)
        updates["item_code"] = item.item_code
        updates["item_name"] = item.item_name

    if "production_date" in payload:
        production_date = _parse_date(payload.get("production_date"))
        if not production_date or production_date != date.today():
            return _build_error("Production date can only be set to today.", 400)
        updates["production_date"] = production_date

    if not updates:
        return _build_error("No valid fields to update.", 400)

    try:
        for key, value in updates.items():
            setattr(entry, key, value)
        db.session.add(entry)
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        return _build_error("Unable to update entry right now.", 500)

    user_lookup = _load_user_lookup({entry.created_by_user_id})
    serial = entry.serials[0] if entry.serials else ExsolProductionSerial(serial_no="00000000")
    return jsonify({"ok": True, "entry": _serialize_entry(entry, serial, user_lookup)})


@bp.get("/template")
@jwt_required()
def download_template():
    if not _has_exsol_production_access():
        return _build_error("Access denied", 403)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exsol Production"
    sheet.append(
        [
            "production_date",
            "item_code",
            "quantity",
            "production_shift",
            "remarks",
            "starting_serial",
            "serial_numbers",
        ]
    )
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="exsol_production_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.post("/excel")
@jwt_required()
def upload_excel():
    if not _has_exsol_production_access():
        return _build_error("Access denied", 403)

    if "file" not in request.files:
        return _build_error("Excel file is required.", 400)
    file = request.files["file"]
    if not file or file.filename == "":
        return _build_error("Excel file is required.", 400)

    try:
        workbook = load_workbook(file, data_only=True)
    except Exception:
        return _build_error("Unable to read the uploaded Excel file.", 400)

    sheet = workbook.active
    try:
        header_cells = next(sheet.iter_rows(max_row=1))
    except StopIteration:
        return _build_error("Excel template headers are missing.", 400)
    header_row = [str(cell.value or "").strip().lower() for cell in header_cells]
    header_map = {header: idx for idx, header in enumerate(header_row) if header}

    required_headers = {"production_date", "item_code", "quantity", "production_shift"}
    if not required_headers.issubset(header_map.keys()):
        return _build_error("Excel template headers are missing.", 400)

    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        payload: dict[str, Any] = {}
        for header, idx in header_map.items():
            payload[header] = row[idx] if idx < len(row) else None
        serial_numbers = payload.get("serial_numbers")
        starting_serial = payload.get("starting_serial")
        if serial_numbers:
            payload["serial_mode"] = "Manual"
            payload["serials"] = serial_numbers
        else:
            payload["serial_mode"] = "SerialRange"
            payload["start_serial"] = starting_serial
        rows.append(payload)

    if not rows:
        return _build_error("Excel file contains no production rows.", 400)

    user_id_raw = get_jwt_identity()
    try:
        user_id = int(user_id_raw)
    except (TypeError, ValueError):
        return _build_error("Invalid user identity.", 403)

    claims = get_jwt()
    role_name = (claims.get("role") or "").strip()
    user = User.query.get(user_id)
    user_name = user.name if user else ""

    try:
        entries, serials_created = _validate_bulk_rows(rows, user_id, role_name, user_name)
    except ExsolProductionValidationError as exc:
        _log_bulk_failure("validation_error", user_id, role_name, rows, errors=exc.errors)
        return jsonify({"ok": False, "errors": exc.errors}), 400
    except SQLAlchemyError:
        _log_bulk_failure("database_error", user_id, role_name, rows)
        return _build_error("Unable to save production entries right now.", 500)

    try:
        _ensure_exsol_sequences()
        for entry in entries:
            db.session.add(entry)
        db.session.flush()
        event_rows = []
        event_created_at = datetime.utcnow()
        for entry in entries:
            event_date = entry.created_at or event_created_at
            for serial in entry.serials:
                event_rows.append(
                    {
                        "company_key": EXSOL_COMPANY_KEY,
                        "item_code": entry.item_code,
                        "serial_number": serial.serial_no,
                        "event_type": "PRODUCED",
                        "event_date": event_date,
                        "ref_type": "PRODUCTION_ENTRY",
                        "ref_id": str(entry.id),
                        "created_at": event_created_at,
                    }
                )
        if not event_rows:
            entry_ids = [entry.id for entry in entries]
            if entry_ids:
                serial_rows = (
                    db.session.query(ExsolProductionSerial, ExsolProductionEntry)
                    .join(ExsolProductionEntry, ExsolProductionSerial.entry_id == ExsolProductionEntry.id)
                    .filter(ExsolProductionSerial.entry_id.in_(entry_ids))
                    .all()
                )
                for serial, entry in serial_rows:
                    event_rows.append(
                        {
                            "company_key": EXSOL_COMPANY_KEY,
                            "item_code": entry.item_code,
                            "serial_number": serial.serial_no,
                            "event_type": "PRODUCED",
                            "event_date": entry.created_at or event_created_at,
                            "ref_type": "PRODUCTION_ENTRY",
                            "ref_id": str(entry.id),
                            "created_at": event_created_at,
                        }
                    )
        if event_rows:
            db.session.execute(insert(ExsolSerialEvent), event_rows)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        _log_bulk_failure("unique_constraint_violation", user_id, role_name, rows, serials=serials_created)
        return _build_error(
            "One or more serials already exist. Please refresh and try again.",
            400,
        )
    except SQLAlchemyError:
        db.session.rollback()
        _log_bulk_failure("database_error", user_id, role_name, rows, serials=serials_created)
        return _build_error("Unable to save production entries right now.", 500)

    summary = {
        "rows_processed": len(entries),
        "total_quantity": sum(entry.quantity for entry in entries),
        "total_serials": len(serials_created),
    }

    return jsonify(
        {
            "ok": True,
            "entry_ids": [entry.id for entry in entries],
            "created_count": len(entries),
            "serials_created": serials_created,
            "summary": summary,
        }
    )
